#!/usr/bin/env python
# -*- coding: utf-8 -*-
import csv
import config
import urllib.request 

from openpyxl import load_workbook


FILE_NAME = 'unterm-efsrca.xlsx'

def run():
    # Setup
    req = urllib.request.Request(config.UNTERM_EFSRCA_URL, headers=config.CONFIG_HEADERS)

    # Download the file and save it
    with urllib.request.urlopen(req) as response, open('/tmp/' + FILE_NAME, 'wb') as out_file:
        out_file.write(response.read())

    wb = load_workbook('/tmp/' + FILE_NAME)
    sheet1 = wb['Worksheet1']
    
    # Write the data to a CSV file
    with open('data/unterm_names.csv', 'w', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow(config.UNTERM_HEADERS)

        # Iterate through the rows of the sheet and write to the CSV
        for row in sheet1.iter_rows(min_row=2, max_col=12, max_row=200):
            values = [cell.value for cell in row]

            if all(values):
                csv_writer.writerow(values)

if __name__ == '__main__':
    run()